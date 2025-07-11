from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import numpy as np
import pdfplumber
import re
import glob
import os
import threading
import tempfile
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

app = Flask(__name__)
progress = {"status": "Not started", "percentage": 0}
results = {"data": None}
download_ready = {"ready": False}

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
warnings.filterwarnings('ignore')

HIGHLIGHTED_EXCEL_PATH = "grouped_lane_mismatches_highlighted.xlsx"

# ------------------------- Helper Functions -------------------------

def update_progress(step, total_steps):
    progress["percentage"] = int((step / total_steps) * 100)
    progress["status"] = f"Processing step {step} of {total_steps}..."

def clean_column_names(columns):
    return [re.sub(r'[^A-Za-z0-9]+', ' ', col) for col in columns]

def read_excel_range(file_path, sheet_name='New Tariff Request Form', start_row=9, start_col=2, end_col=46):
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl', header=None)
    df_range = df.iloc[start_row - 1:, start_col - 1:end_col].dropna(how='all')
    df_range.columns = df_range.iloc[1]
    df_range = df_range.drop(df_range.index[:4])
    df_range.columns = clean_column_names(df_range.columns)
    return df_range.reset_index(drop=True)

def extract_tables_except_last(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        tables = []
        for i in range(len(pdf.pages) - 1):
            page = pdf.pages[i]
            table = page.extract_table()
            if table:
                tables += table
    return tables

def read_small_boxes(pdf_path, page_number, box_coordinates):
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[page_number - 1]
        boxes_content = []
        for box_coords in box_coordinates:
            x0, y0, x1, y1 = box_coords
            box = page.crop((x0, y0, x1, y1))
            box_text = box.extract_tables()
            boxes_content.append(box_text)
    return boxes_content

def read_content(tables):
    content_table = []
    for table in tables:
        for tab in table:
            content_table.append(tab)
    return content_table

def highlight_mismatches_in_excel(csv_file, output_excel_file):
    df = pd.read_csv(csv_file)
    grouped = df.groupby(['Origin', 'Destination'])

    rows_to_write = []
    for (origin, destination), group in grouped:
        rows_to_write.append({
            'Source': '',
            'Origin': f'Lane: {origin} ➔ {destination}',
            'Destination': '',
            'Rate': '',
            'Mismatch Detail': ''
        })
        rows_to_write.extend(group.to_dict(orient='records'))
        rows_to_write.append({'Source': '', 'Origin': '', 'Destination': '', 'Rate': '', 'Mismatch Detail': ''})

    grouped_df = pd.DataFrame(rows_to_write)
    grouped_df.to_excel(output_excel_file, index=False)

    wb = load_workbook(output_excel_file)
    ws = wb.active

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold_font = Font(bold=True)

    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        mismatch_value = row[headers['Mismatch Detail'] - 1].value if 'Mismatch Detail' in headers else None
        origin_value = row[headers['Origin'] - 1].value if 'Origin' in headers else ''

        if mismatch_value and 'Match' not in mismatch_value:
            for cell in row:
                cell.fill = red_fill
            row[headers['Mismatch Detail'] - 1].font = bold_font

        if isinstance(origin_value, str) and origin_value.startswith("Lane:"):
            for cell in row:
                cell.font = Font(bold=True)

    wb.save(output_excel_file)

# ------------------------- Web Routes -------------------------

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/progress')
def get_progress():
    return jsonify(progress)

@app.route('/download_status')
def download_status():
    return jsonify({"ready": download_ready["ready"]})

@app.route('/download')
def download_file():
    if os.path.exists(HIGHLIGHTED_EXCEL_PATH):
        return send_file(HIGHLIGHTED_EXCEL_PATH, as_attachment=True)
    return "File not found", 404

@app.route('/process', methods=['POST'])
def process():
    excel_files = request.files.getlist('excel_files')
    pdf_files = request.files.getlist('pdf_files')

    if not excel_files or not pdf_files:
        return jsonify({"error": "No files uploaded"}), 400

    excel_bytes = excel_files[0].read()
    pdf_file_data = [(f.filename, f.read()) for f in pdf_files]
    download_ready["ready"] = False

    def worker(excel_bytes, pdf_file_data):
        try:
            update_progress(1, 5)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_excel:
                temp_excel.write(excel_bytes)
                temp_excel_path = temp_excel.name

            final_excel_df = read_excel_range(temp_excel_path)

            update_progress(2, 5)
            all_pdf_df = pd.DataFrame()
            box_coordinates = [(0, 70, 1400, 420)]

            for filename, pdf_bytes in pdf_file_data:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
                    temp_pdf.write(pdf_bytes)
                    temp_pdf_path = temp_pdf.name

                final_ori_table = extract_tables_except_last(temp_pdf_path)
                if not final_ori_table:
                    continue
                ori_df = pd.DataFrame(final_ori_table[2:], columns=final_ori_table[1])
                strings_to_remove = ['Carrier Legal Name:', 'Lane ID', 'Weekly \nLoads \nEstimate']
                filt_ori_df = ori_df[~ori_df['Lane ID'].isin(strings_to_remove)]
                filt_ori_df = filt_ori_df.drop(columns='Weekly \nLoads \nEstimate', errors='ignore')
                small_ori_table = read_small_boxes(temp_pdf_path, 1, box_coordinates)
                if small_ori_table and len(small_ori_table) > 0:
                    output_table = read_content(small_ori_table[0])
                    data_dict = {item[0].strip(':'): item[1] for item in output_table}
                    for key in data_dict:
                        filt_ori_df[key] = data_dict[key]
                all_pdf_df = pd.concat([all_pdf_df, filt_ori_df], ignore_index=True)

            update_progress(3, 5)
            def normalize_text(s):
                return re.sub(' +', ' ', s.replace('\n', ' ')).strip() if isinstance(s, str) else s

            final_excel_df['Origin Detail'] = final_excel_df['Origin Detail'].apply(normalize_text)
            final_excel_df['Destination Detail'] = final_excel_df['Destination Detail'].apply(normalize_text)
            final_excel_df['Lane Currency'] = final_excel_df[' Lane Currency'].apply(normalize_text)
            final_excel_df['Origin Prov'] = final_excel_df[' Origin Prov'].apply(normalize_text)
            final_excel_df['Destination Prov'] = final_excel_df['Destination Prov'].apply(normalize_text)
            final_excel_df['Inbound/Outbound'] = final_excel_df[' Inbound Outbound'].apply(normalize_text)
            final_excel_df['Excel Rate'] = pd.to_numeric(final_excel_df[' Current Rate Per Unit to upload'], errors='coerce')

            all_pdf_df['Rate Origin'] = all_pdf_df['Rate Origin'].apply(normalize_text)
            all_pdf_df['Rate Destination'] = all_pdf_df['Rate Destination'].apply(normalize_text)
            all_pdf_df['Lane Currency'] = all_pdf_df['Lane Currency'].apply(normalize_text)
            all_pdf_df['Origin Prov'] = all_pdf_df['Origin Prov.'].apply(normalize_text)
            all_pdf_df['Destination Prov'] = all_pdf_df['Destination \nProv.'].apply(normalize_text)
            all_pdf_df['Inbound/Outbound'] = all_pdf_df['Inbound/ \nOutbound'].apply(normalize_text)
            all_pdf_df['PDF Rate'] = pd.to_numeric(all_pdf_df['Year 1'].replace('[\$,]', '', regex=True), errors='coerce')

            update_progress(4, 5)
            excel_df = final_excel_df[['Origin Detail', 'Destination Detail', 'Lane Currency',
                                       'Origin Prov', 'Destination Prov', 'Inbound/Outbound', 'Excel Rate']].copy()
            excel_df['Source'] = 'Excel'
            excel_df['Lane'] = list(zip(excel_df['Origin Detail'], excel_df['Destination Detail']))

            pdf_df = all_pdf_df[['Rate Origin', 'Rate Destination', 'Lane Currency',
                                 'Origin Prov', 'Destination Prov', 'Inbound/Outbound', 'PDF Rate']].copy()
            pdf_df.columns = ['Origin Detail', 'Destination Detail', 'Lane Currency',
                              'Origin Prov', 'Destination Prov', 'Inbound/Outbound', 'PDF Rate']
            pdf_df['Source'] = 'PDF'
            pdf_df['Lane'] = list(zip(pdf_df['Origin Detail'], pdf_df['Destination Detail']))

            excel_lanes = set(excel_df['Lane'])
            pdf_lanes = set(pdf_df['Lane'])

            mismatched_lanes = excel_lanes.symmetric_difference(pdf_lanes)
            grouped_rows = []

            for lane in mismatched_lanes:
                origin, destination = lane

                for _, row in excel_df[excel_df['Lane'] == lane].iterrows():
                    grouped_rows.append({
                        'Source': 'Excel',
                        'Origin': origin,
                        'Destination': destination,
                        'Lane Currency': row['Lane Currency'],
                        'Origin Prov': row['Origin Prov'],
                        'Destination Prov': row['Destination Prov'],
                        'Inbound/Outbound': row['Inbound/Outbound'],
                        'Rate': row['Excel Rate'],
                        'Mismatch Detail': 'No match in PDF'
                    })

                for _, row in pdf_df[pdf_df['Lane'] == lane].iterrows():
                    grouped_rows.append({
                        'Source': 'PDF',
                        'Origin': origin,
                        'Destination': destination,
                        'Lane Currency': row['Lane Currency'],
                        'Origin Prov': row['Origin Prov'],
                        'Destination Prov': row['Destination Prov'],
                        'Inbound/Outbound': row['Inbound/Outbound'],
                        'Rate': row['PDF Rate'],
                        'Mismatch Detail': 'No match in Excel'
                    })

            mismatch_df = pd.DataFrame(grouped_rows)
            mismatch_df.to_csv("grouped_lane_mismatches.csv", index=False)
            highlight_mismatches_in_excel("grouped_lane_mismatches.csv", HIGHLIGHTED_EXCEL_PATH)
            results['data'] = mismatch_df.to_html(classes='data')
            download_ready["ready"] = True
            update_progress(5, 5)
            progress["status"] = "Completed"
        except Exception as e:
            progress["status"] = f"Error: {str(e)}"

    thread = threading.Thread(target=worker, args=(excel_bytes, pdf_file_data))
    thread.start()
    return jsonify({"message": "Processing started"})

@app.route('/results')
def show_results():
    return render_template('results.html', table=results.get("data"))


# ====== NEW API ENDPOINTS =======

@app.route('/api/summary')
def api_summary():
    if not os.path.exists(HIGHLIGHTED_EXCEL_PATH):
        return jsonify({"error": "Summary not available"}), 404

    df = pd.read_excel(HIGHLIGHTED_EXCEL_PATH)
    total_rows = len(df)
    mismatches = len(df[df.get('Mismatch Detail') != 'Match']) if 'Mismatch Detail' in df.columns else 0
    unique_lanes = df['Origin'].nunique() if 'Origin' in df.columns else 0

    summary = {
        "total_rows": total_rows,
        "mismatches": mismatches,
        "unique_lanes": unique_lanes
    }

    return jsonify(summary)

@app.route('/api/results')
def api_results():
    draw = int(request.args.get('draw', 1))
    start = int(request.args.get('start', 0))
    length = int(request.args.get('length', 10))
    search_value = request.args.get('search[value]', '').lower()

    if not os.path.exists(HIGHLIGHTED_EXCEL_PATH):
        return jsonify({
            "draw": draw,
            "recordsTotal": 0,
            "recordsFiltered": 0,
            "data": []
        })

    df = pd.read_excel(HIGHLIGHTED_EXCEL_PATH)
    df.fillna('', inplace=True)

    records_total = len(df)

    if search_value:
        df = df[df.apply(lambda row: row.astype(str).str.lower().str.contains(search_value).any(), axis=1)]

    records_filtered = len(df)
    paginated_df = df.iloc[start:start+length]

    return jsonify({
        "draw": draw,
        "recordsTotal": records_total,
        "recordsFiltered": records_filtered,
        "data": paginated_df.to_dict(orient='records')
    })


if __name__ == '__main__':
    app.run(debug=True)
